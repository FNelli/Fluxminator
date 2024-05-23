import os
import openpyxl
from openpyxl import load_workbook

import time
from datetime import datetime
from datetime import timedelta

from fluxminator.Constants import *
from fluxminator.Result import Result


class Runner:
    """ This object is responsible for executing the simulations; it is inherited by the Presenter class.
    The main reason behind it is to separate the scripts for easier understanding and future maintenance. """

    # Workbooks:

    summary_wb = None
    summary_worksheets = [[], [], []]                               # stack(s), gammaDep, loadDep

    cogging_wb = None
    cogging_fft_wb = None

    benf_wb = None
    bemf_fft_wb = None

    ripple_wb = None
    ripple_fft_wb = None

    stack1 = False
    stack2 = False
    stack3 = False

    @staticmethod
    def _run(presenter):

        time.sleep(5)

        # Try to load the already existing summary and detailed result files
        if presenter.model.is_sessionOld() or presenter.model.get_simulationID() > 1:
            if Runner._load_workbooks(presenter) == RESULT_FILE_IO_ERROR:
                return RESULT_FILE_IO_ERROR

        simulation_index = presenter.model.get_simulationID()

        try:
            wb_sets = openpyxl.load_workbook(presenter.model.fluxModel.motor.get_parameterSetExcelPath())
            ws_sets = wb_sets.worksheets[0]
            sets_row_number = ws_sets.max_row                       # 3 header rows + parameter set rows with id-s
            sets_col_number = ws_sets.max_column
            sets_number = sets_row_number - 3                       # first 3 rows -> param. type, name, description

            # COMMON PARAMETER DATA FOR THE CHANNEL FILE (Types & Names):

            parameter_types_row = ws_sets[1]
            parameter_types_list = [p.value for p in parameter_types_row][1:]
            parameter_types_string = " ".join(parameter_types_list)

            parameter_names_row = ws_sets[2]
            parameter_names_list = [p.value for p in parameter_names_row][1:]
            parameter_names_string = " ".join(parameter_names_list)

            parameter_desc_row = ws_sets[3]                         # this one is only for result file creation
            parameter_desc_list = [p.value for p in parameter_desc_row][1:]

            parameter_string = parameter_types_string + "\n" + parameter_names_string + "\n{values}"
            channel_data_string = presenter.model.get_channelData().format(id="{id}", params=parameter_string)

            # PROGRESS BAR INITIALIZATION FOR THE EXECUTION:

            presenter.view.progressBar.setMaximum(sets_number)
            Runner._set_progress(presenter=presenter, value=(presenter.model.get_simulationID() - 1))

            stop_the_summary_creator = False                        # stop when there are no more result files available

            # MAIN LOOP: one iteration for each parameter combination (simulation in Flux)
            while simulation_index < sets_number + 1:

                if not presenter.model.get_executionInProgress():               # the STOP button was pressed
                    break

                print("-------------------------")
                print("SIMULATION ID:", sets_number, "/", simulation_index)
                print("-------------------------")

                Runner._update_estimated_time(presenter=presenter, index=simulation_index, sets_number=sets_number)

                parameter_values_row = ws_sets[simulation_index + 3]            # 3 header rows in the Excel file
                parameter_values_id = parameter_values_row[0].value
                parameter_values_list = [p.value for p in parameter_values_row][1:]

                # Search for already existing simulation result files from Flux:
                if not presenter.model.is_sessionNew() or presenter.model.get_summarySavingProblem():

                    if presenter.model.get_summarySavingProblem():
                        presenter.model.set_summarySavingProblem(False)         # reset this variable, try to save again

                    simulation_is_needed = not Runner._check_existing_result_files(presenter=presenter,
                                                                                   row_id=parameter_values_id)
                    print("Simulation is needed:", simulation_is_needed)

                    # When simulation is required for summary creation in the summary session, we can't continue:
                    if simulation_is_needed and presenter.model.is_sessionSummary():
                        if simulation_index == 1:
                            return MISSING_RAW_FILES                            # no raw results files at all
                        stop_the_summary_creator = True
                else:
                    simulation_is_needed = True

                # FINALLY - START A NEW SIMULATION IN  FLUX:
                if simulation_is_needed and not stop_the_summary_creator:
                    Runner._execute_flux_simulation(presenter, channel_data_string,
                                                    parameter_values_list, parameter_values_id)

                # PROCESS RAW DATA FILES:
                if not stop_the_summary_creator:

                    result = Result()

                    file_path = presenter.model.fluxModel.get_modelPath() + "/Results/"
                    file_name = presenter.model.fluxModel.get_modelName()[:-4] + "_" + parameter_values_id

                    # Create Cogging & BEMF objects:
                    if presenter.model.is_scenarioCogging():
                        file = file_path + file_name + "_Cogging_BEMF.xls"
                        if not result.setup_Cogging_BEMF(file, presenter.model.fluxModel, presenter.model.get_scenarioID):
                            return RAW_COGGING_FILE_ERROR

                    # Create Ripple objects:
                    if presenter.model.is_scenarioRipple():
                        file = file_path + file_name + "_LOAD_{0}_GAMMA_{1}.xls"
                        if not result.setup_Ripple(file=file, flux_model=presenter.model.fluxModel):
                            return RAW_RIPPLE_FILE_ERROR

                # CREATE PRETTY EXCEL FILES:

                try:
                    if not stop_the_summary_creator:
                        Runner._create_summary_file(presenter, result, simulation_index, parameter_values_id,
                                                    parameter_desc_list, parameter_names_list, parameter_values_list)

                        # TODO DETAILED FILES - FUTURE DEVELOPMENT

                    # Saving in summary mode: only at the end (stop_the_summary_creator = True)
                    # Saving in normal mode:  after every simulation
                    if stop_the_summary_creator or not presenter.model.is_sessionSummary() or \
                            simulation_index == sets_number:

                        path = presenter.model.fluxModel.get_modelPath() + "/Results/{}.xlsx"

                        Runner.summary_wb.save(path.format("Summary"))

                except IOError:
                    presenter.model.set_simulationID(simulation_index)
                    return RESULT_FILE_IO_ERROR

                if not stop_the_summary_creator:
                    simulation_index += 1
                else:
                    break

            wb_sets.close()

            if simulation_index > sets_number:
                presenter.model.set_simulationID(1)
                return EXECUTION_DONE

            elif presenter.model.is_sessionSummary():
                presenter.model.set_simulationID(simulation_index)
                return EXECUTION_DONE

            else:
                presenter.model.set_simulationID(simulation_index)
                return EXECUTION_STOP

        except:
            presenter.model.set_simulationID(simulation_index)
            return EXECUTION_ERROR

    @staticmethod
    def _load_workbooks(presenter):
        """ Try to load the already existing summary and detailed result files """

        if not (os.path.exists(presenter.model.fluxModel.get_modelPath() + "/Results/Summary.xlsx")):
            presenter.model.set_simulationID(1)
        else:
            try:
                # SUMMARY FILE:

                Runner.summary_wb = load_workbook(presenter.model.fluxModel.get_modelPath() + "/Results/Summary.xlsx")
                sheets = Runner.summary_wb.sheetnames

                # Search for valid sheets in the summary Excel:
                if not ("1 Stack" in sheets or "2 Stacks" in sheets or "3 Stacks" in sheets):
                    presenter.model.set_simulationID(1)
                else:
                    # Search for the next ID (check every sheet, choose the smallest number (but there should be equal))
                    next_simulation_id = []
                    if "1 Stack" in sheets:
                        Runner.stack1 = True
                        ws_summary1 = Runner.summary_wb.get_sheet_by_name("1 Stack")
                        next_simulation_id.append(ws_summary1.max_row - 1)
                        Runner.summary_worksheets[0].append([0, ws_summary1])

                    if "2 Stacks" in sheets:
                        Runner.stack2 = True
                        ws_summary2 = Runner.summary_wb.get_sheet_by_name("2 Stacks")
                        next_simulation_id.append(ws_summary2.max_row - 1)
                        Runner.summary_worksheets[0].append([1, ws_summary2])

                    if "3 Stacks" in sheets:
                        Runner.stack3 = True
                        ws_summary3 = Runner.summary_wb.get_sheet_by_name("3 Stacks")
                        next_simulation_id.append(ws_summary3.max_row - 1)
                        Runner.summary_worksheets[0].append([2, ws_summary3])

                    presenter.model.set_simulationID(min(next_simulation_id))

                if presenter.model.is_dependencyGamma():
                    if "Gamma Dependency" not in sheets:
                        presenter.model.set_simulationID(1)
                    else:
                        Runner.summary_worksheets[1] = Runner.summary_wb.get_sheet_by_name("Gamma Dependency")

                if presenter.model.is_dependencyLoad():
                    if "Load Dependency" not in sheets:
                        presenter.model.set_simulationID(1)
                    else:
                        Runner.summary_worksheets[2] = Runner.summary_wb.get_sheet_by_name("Load Dependency")

                # DETAILED FILES: TODO
                """
                path_template = presenter.model.fluxModel.get_modelPath() + "/Results/Detailed_{}.xlsx"

                if presenter.model.get_simulationID() != 1 and presenter.model.is_scenarioCogging():
                    if os.path.exists(path_template.format("Cogging_Values")) and \
                       os.path.exists(path_template.format("Cogging_FFT")) and \
                       os.path.exists(path_template.format("BEMF_Values")) and \
                       os.path.exists(path_template.format("BEMF_FFT")):

                        Runner.cogging_wb = load_workbook(path_template.format("Cogging_Values"))
                        Runner.cogging_fft_wb = load_workbook(path_template.format("Cogging_FFT"))
                        Runner.bemf_wb = load_workbook(path_template.format("BEMF_Values"))
                        Runner.bemf_fft_wb = load_workbook(path_template.format("BEMF_FFT"))
                    else:
                        presenter.model.set_simulationID(1)

                if presenter.model.get_simulationID() != 1 and presenter.model.is_scenarioRipple():
                    if os.path.exists(path_template.format("Ripple_Values")) and \
                       os.path.exists(path_template.format("Ripple_FFT")):

                        Runner.ripple_wb = load_workbook(path_template.format("Ripple_Values"))
                        Runner.ripple_fft_wb = load_workbook(path_template.format("Ripple_FFT"))
                    else:
                        presenter.model.set_simulationID(1)
                """
            except IOError:
                return RESULT_FILE_IO_ERROR

    @staticmethod
    def _check_existing_result_files(presenter, row_id):
        """ Check if the raw data files for the current parameter value combination already exist """

        # COGGING SCENARIO:
        if presenter.model.is_scenarioCogging():
            path_to_check = presenter.model.fluxModel.get_modelPath() + "/Results/" \
                            + presenter.model.fluxModel.get_modelName()[:-4] + "_" + row_id + "_Cogging_BEMF.xls"
            if not os.path.exists(path_to_check):
                return False

        # RIPPLE SCENARIO:
        if presenter.model.is_scenarioRipple():
            path_template = presenter.model.fluxModel.get_modelPath() + "/Results/" \
                            + presenter.model.fluxModel.get_modelName()[:-4] + "_" + row_id + "_LOAD_{0}_GAMMA_{1}.xls"

            current_values = presenter.model.fluxModel.rangeCurrent.get_all_values()
            gamma_values = presenter.model.fluxModel.rangeGamma.get_all_values()

            for cu_index in range(1, len(current_values) + 1):
                for g_index in range(1, len(gamma_values) + 1):

                    if not os.path.exists(path_template.format(cu_index, g_index)):
                        return False

        # we reach this point -> all the necessary files are existing
        return True

    @staticmethod
    def _execute_flux_simulation(presenter, channel_data_string, parameter_values_list, parameter_values_id):

        # CHANNEL FILE:
        parameter_values_string = ""
        for value in parameter_values_list:
            parameter_values_string += "%.2f" % value + " "

        upcoming_channel_data = channel_data_string.format(id=parameter_values_id,
                                                           values=parameter_values_string)

        channel_file = open("Channel.txt", 'w')  # update Channel file for Flux
        channel_file.write(upcoming_channel_data)
        channel_file.close()

        # RUN FLUX:
        start_time = time.time()  # measure simulation time

        cwd = os.getcwd()
        os.system('cd ' + presenter.model.fluxModel.get_modelPath())
        flux_path = presenter.flux_application_path + FLUX_EXE_PATH
        run_flux_command = '"' + flux_path + '" -runPy ' + \
                           str(cwd) + '\\' + PYFLUX_SCRIPT_NAME + ' -application Flux2D -batch'
        os.system(run_flux_command)     # os.system() method executes the command (a string) in a subshell

        end_time = time.time()
        presenter.model.extend_simulationTimes(int(end_time - start_time))

    @staticmethod
    def _set_progress(presenter, reset=False, value=None):
        """ Set the progress state in the model and display it on the view (reset is called by the Presenter class) """

        if reset:
            presenter.model.set_progressState(0)
            presenter.model.simulationTimes = []
            presenter.view.progressLabel.setText("Calculation: 0/0, estimated finish time: -")
            presenter.view.progressBar.setValue(0)

        elif value is not None:
            presenter.model.set_progressState(value + 1)            # the number of the next simulation
            presenter.view.progressBar.setValue(value)              # the number of simulations already executed

        else:
            progress_value = presenter.model.get_progressState()
            presenter.model.set_progressState(progress_value + 1)
            presenter.view.progressBar.setValue(progress_value)

    @staticmethod
    def _update_estimated_time(presenter, index, sets_number):
        """ Update the estimated finish time of the execution based on the data from previous simulation executions """

        if len(presenter.model.get_simulationTimes()) == 0:
            presenter.view.progressLabel.setText("Calculation: {0}/{1}, "
                                                 "estimated finish time: {2}".format(index, sets_number, "-"))
            return

        time_in_sec = 0
        for time_data in presenter.model.get_simulationTimes():
            time_in_sec += time_data
        mean_time_in_sec = time_in_sec / len(presenter.model.get_simulationTimes())

        remaining_time_in_sec = mean_time_in_sec * (sets_number - index + 1)

        current_date = datetime.now() + timedelta(seconds=remaining_time_in_sec)

        presenter.view.progressLabel.setText("Calculation: {0}/{1}, estimated finish time: {2}".format(
            index, sets_number, current_date.strftime("%b. %d (%A) %H:%M")))

    @staticmethod
    def _create_summary_file(presenter, result, simulation_index, parameter_values_id, parameter_desc_list,
                             parameter_names_list, parameter_values_list):

        if simulation_index == 1:                                   # no data in Summary file yet
            # INITIALIZE WORKBOOK:
            Runner._initialize_summary_file(presenter, parameter_names_list, parameter_desc_list, result)

        # PARAMETER VALUES:
        common_data_for_stack_sheets = parameter_values_list + [parameter_values_id]

        # COGGING & BEMF RELATED DATA:
        cogging_p2p, cogging_amplitudes, cogging_phases, bemf_ph_rms, bemf_amplitudes, bemf_phases = \
            Runner._get_cogging_bemf_data(presenter, result)

        # RIPPLE RELATED DATA:
        main_parameters_data, ripple_amplitudes, ripple_phases = \
            Runner._get_current_dependent_data(presenter, result, cogging_p2p)

        for ws_container in Runner.summary_worksheets[0]:   # ws_container = [index, worksheet]

            ws_index = ws_container[0]

            actual_data = common_data_for_stack_sheets + [cogging_p2p[ws_index]] + main_parameters_data[ws_index] + \
                [bemf_ph_rms[ws_index]] + cogging_amplitudes[ws_index] + ripple_amplitudes[ws_index] + \
                bemf_amplitudes[ws_index] + cogging_phases[ws_index] + ripple_phases[ws_index] + bemf_phases[ws_index]

            ws_container[1].append(actual_data)

    @staticmethod
    def _initialize_summary_file(presenter, parameter_names_list, parameter_desc_list, result):
        """ Create relevant worksheets for the summary file and call header creator functions """

        # SET UP THE HEADER ROWS:
        header_row1, header_row2 = Runner._create_summary_file_header(presenter, parameter_names_list,
                                                                      parameter_desc_list, result)
        initialized_workbook = False

        # EXCEL SHEETS FOR 1/2/3 STACKS:
        if result.get_cogging1Stack() is not None or result.get_ripple1Stack() != {}:
            Runner.stack1 = True

            Runner.summary_wb = Workbook()
            initialized_workbook = True
            ws_summary1 = Runner.summary_wb.active
            ws_summary1.title = "1 Stack"

            ws_summary1.append(header_row1)
            ws_summary1.append(header_row2)

            Runner.summary_worksheets[0].append([0, ws_summary1])   # worksheets in array -> easy to handle with loops

        if result.get_cogging2Stack() is not None or result.get_ripple2Stack() != {}:
            Runner.stack2 = True

            if not initialized_workbook:
                Runner.summary_wb = Workbook()
                initialized_workbook = True
                ws_summary2 = Runner.summary_wb.active
                ws_summary2.title = "2 Stacks"
            else:
                ws_summary2 = Runner.summary_wb.create_sheet("2 Stacks")

            ws_summary2.append(header_row1)
            ws_summary2.append(header_row2)

            Runner.summary_worksheets[0].append([1, ws_summary2])

        if result.get_cogging3Stack() is not None or result.get_ripple3Stack() != {}:
            Runner.stack3 = True

            if not initialized_workbook:
                Runner.summary_wb = Workbook()
                initialized_workbook = True
                ws_summary3 = Runner.summary_wb.active
                ws_summary3.title = "3 Stacks"
            else:
                ws_summary3 = Runner.summary_wb.create_sheet("3 Stacks")

            ws_summary3.append(header_row1)
            ws_summary3.append(header_row2)

            Runner.summary_worksheets[0].append([2, ws_summary3])

        # EXCEL SHEETS FOR GAMMA & LOAD DEPENDENCY: TODO FUTURE DEVELOPMENT
        """
        if presenter.model.is_dependencyGamma():
            ws_summary_gamma = Runner.summary_wb.create_sheet("Gamma Dependency")
            gamma_header_row = ["", "", "Gamma"]

            for gamma_value in presenter.model.rangeGamma.get_all_values():
                gamma_header_row.append(gamma_value)

            Runner.summary_worksheets[1] = ws_summary_gamma

        if presenter.model.is_dependencyLoad():
            ws_summary_load = Runner.summary_wb.create_sheet("Load Dependency")
            load_header_row = ["", "", "Load [A]"]

            for current_value in presenter.model.rangeCurrent.get_all_values():
                load_header_row.append(current_value)

            Runner.summary_worksheets[2] = ws_summary_load
        """
    @staticmethod
    def _create_summary_file_header(presenter, parameter_names_list, parameter_desc_list, result):

        current_values = presenter.model.fluxModel.rangeCurrent.get_all_values()

        header_row1 = parameter_names_list + ["", ""]
        header_row2 = parameter_desc_list + ["No.", "Cogging\np2p\n[Ncm]"]

        for current_value in current_values:
            header_row1 += ["Load = {} A".format(current_value), "", "", ""]
            header_row2 += ["Cogging\np2p [%]", "T_mean\n[Nm]", "Ripple\np2p\n[Ncm]", "Ripple\np2p [%]"]

        header_row1 += [""]
        header_row2 += ["BackEMF\nph RMS\n[V]"]

        bemf_index = len(header_row2)                           # cell index for the red background color (if invalid)

        amplitude_phase_strings = [["Ampl.", "[Ncm]", "Ampl. [V]"], ["Phase", "[rad]", "Phase [rad]"]]

        for strings in amplitude_phase_strings:

            # COGGING AMPLITUDES / PHASES:
            if not presenter.model.is_scenarioCogging():        # no Cogging scenario
                for _ in range(20):
                    header_row1.append("")
                    header_row2.append("Cogging\n-\n{0}\n{1}".format(strings[0], strings[1]))

            else:
                cogging = result.get_cogging1Stack()            # for the header it doesn't matter which one
                if cogging is None:
                    cogging = result.get_cogging2Stack()
                if cogging is None:
                    cogging = result.get_cogging3Stack()

                i = 0
                for _ in range(1, len(cogging.get_fftValues()[0])):         # we need the harmonics
                    i += 1                                                  # but not the 0. harmonic
                    header_row1.append("")
                    header_row2.append("Cogging\n{2}th\n{0}\n{1}".format(strings[0], strings[1],
                                                                         int(cogging.get_fftValues()[0][i])))
                    if i == 20:
                        break  # we need max 20 harmonics

                while i < 20:
                    header_row1.append("")
                    header_row2.append("Cogging\n-\n{0}\n{1}".format(strings[0], strings[1]))
                    i += 1

            # RIPPLE AMPLITUDES / PHASES:
            if not presenter.model.is_scenarioRipple():                     # no Ripple scenario
                for j in range(20):
                    if j == 0:
                        header_row1.append("Load = 0 A")
                    else:
                        header_row1.append("")
                    header_row2.append("Ripple\n-\n{0}\n{1}".format(strings[0], strings[1]))
            else:
                ripple = result.get_rippleModels()[current_values[0]][0]    # for the header it doesn't matter which one

                for current_value in current_values:
                    j = 0
                    for _ in range(len(ripple.get_fftValues()[0])):         # we need the harmonics
                        j += 1                                              # but not the 0. harmonic

                        if j == 1:
                            header_row1.append("Load = {} A".format(current_value))
                        else:
                            header_row1.append("")

                        header_row2.append("Ripple\n{2}th\n{0}\n{1}".format(strings[0], strings[1],
                                                                            int(ripple.get_fftValues()[0][j])))
                        if j == 20:
                            break

                    while j < 20:
                        header_row1.append("")
                        header_row2.append("Ripple\n-\n{0}\n{1}".format(strings[0], strings[1]))
                        j += 1

            # BEMF AMPLITUDES / PHASES:
            if not presenter.model.is_scenarioCogging():                    # no BEMF
                for _ in range(20):
                    header_row1.append("")
                    header_row2.append("ph\nBackEMF\n-\n{0}".format(strings[2]))

            else:
                bemf = result.get_BEMF1Stack()                           # for the header it doesn't matter which one
                if bemf is None:
                    bemf = result.get_BEMF2Stack()
                if bemf is None:
                    bemf = result.get_BEMF3Stack()

                if bemf.is_valid_BEMF():
                    k = 0
                    for _ in range(1, len(bemf.get_fftValues()[0])):
                        k += 1
                        header_row1.append("")
                        header_row2.append("ph\nBackEMF\n{1}\n{0}".format(strings[2], int(bemf.fftValues[0][j])))

                        if k == 20:
                            break

                    while k < 20:
                        header_row1.append("")
                        header_row2.append("ph\nBackEMF\n-\n{0}".format(strings[2]))
                        k += 1
                else:
                    for _ in range(20):
                        header_row1.append("")
                        header_row2.append("ph\nBackEMF\n-\n{0}".format(strings[2]))

        return header_row1, header_row2

    @staticmethod
    def _get_cogging_bemf_data(presenter, result):

        stack_booleans = [Runner.stack1, Runner.stack2, Runner.stack3]
        stack_coggings = [result.get_cogging1Stack(), result.get_cogging2Stack(), result.get_cogging3Stack()]
        stack_bemfs = [result.get_BEMF1Stack(), result.get_BEMF2Stack(), result.get_BEMF3Stack()]

        # cogging & bemf:
        cogging_p2p = []
        cogging_amplitudes = [[], [], []]                       # placeholders for 1/2/3 Stack(s) data
        cogging_phases = [[], [], []]

        bemf_ph_rms = []
        bemf_amplitudes = [[], [], []]
        bemf_phases = [[], [], []]

        c = 0                                                   # counter for the arrays above
        for stack in stack_booleans:

            if stack:

                if not presenter.model.is_scenarioCogging():                                # no cogging scenario
                    cogging_p2p.append("")
                    bemf_ph_rms.append("")
                    for _ in range(20):
                        cogging_amplitudes[c].append("")
                        cogging_phases[c].append("")
                        bemf_amplitudes[c].append("")
                        bemf_phases[c].append("")

                else:                                                                       # cogging scenario
                    # COGGING:
                    cogging_p2p.append(stack_coggings[c].get_p2p())
                    cogging_amplitudes[c] = stack_coggings[c].get_fftValues()[1][1:]        # from 1. harmonic
                    cogging_phases[c] = stack_coggings[c].get_fftValues()[2][1:]

                    if len(cogging_amplitudes[c]) < 20:
                        for _ in range(len(cogging_amplitudes[c]), 20):
                            cogging_amplitudes[c].append("-")
                            cogging_phases[c].append("-")
                    elif len(cogging_amplitudes[c]) > 20:                                   # max 20 harmonics
                        cogging_amplitudes[c] = cogging_amplitudes[c][:20]
                        cogging_phases[c] = cogging_phases[c][:20]

                    # BEMF:
                    if stack_bemfs[c].is_valid_BEMF():
                        bemf_ph_rms.append(stack_bemfs[c].get_rmsValue())
                        bemf_amplitudes[c] = stack_bemfs[c].get_fftValues()[1][1:]          # from 1. harmonic
                        bemf_phases[c] = stack_bemfs[c].get_fftValues()[2][1:]

                        if len(bemf_amplitudes[c]) < 20:
                            for _ in range(len(bemf_amplitudes[c]), 20):
                                bemf_amplitudes[c].append("-")
                                bemf_phases[c].append("-")
                        elif len(bemf_amplitudes[c]) > 20:
                            bemf_amplitudes[c] = bemf_amplitudes[c][:20]
                            bemf_phases[c] = bemf_phases[c][:20]

                    else:
                        bemf_ph_rms.append("")
                        for _ in range(20):
                            bemf_amplitudes[c].append("")
                            bemf_phases[c].append("")

            else:                                               # this skewed type doesn't exist
                cogging_p2p.append(None)
                bemf_ph_rms.append(None)
                cogging_amplitudes[c] = None
                cogging_phases[c] = None
                bemf_amplitudes[c] = None
                bemf_phases[c] = None

            c += 1

        return cogging_p2p, cogging_amplitudes, cogging_phases, bemf_ph_rms, bemf_amplitudes, bemf_phases

    @staticmethod
    def _get_current_dependent_data(presenter, result, cogging_p2p):

        # Empty arrays for 1/2/3 Stack (cogging_p2p is an array with 3 elements):
        main_params = [[], [], []]              # cog. p2p%, torque mean, ripple p2p, ripple p2p%
        ripple_amplitudes = [[], [], []]
        ripple_phases = [[], [], []]

        if not presenter.model.is_scenarioRipple():                     # no ripple scenario
            for i in range(3):
                main_params[i] = ["", "", "", ""]

                for _ in range(20):
                    ripple_amplitudes[i].append("")
                    ripple_phases[i].append("")

        else:
            stack_booleans = [Runner.stack1, Runner.stack2, Runner.stack3]
            stack_ripples = [result.get_ripple1Stack(), result.get_ripple2Stack(), result.get_ripple3Stack()]

            for current_value in presenter.model.fluxModel.rangeCurrent.get_all_values():

                for i in range(3):
                    if stack_booleans[i]:

                        ripple_p2p = stack_ripples[i][current_value].get_p2p()
                        torque_mean = stack_ripples[i][current_value].get_torqueMean()

                        amplitudes = stack_ripples[i][current_value].get_fftValues()[1][1:]         # from 1. harmonic
                        phases = stack_ripples[i][current_value].get_fftValues()[2][1:]

                        ripple_p2p_percent = ripple_p2p / torque_mean

                        if cogging_p2p[i] == "":
                            cogging_p2p_percent = ""
                        else:
                            cogging_p2p_percent = cogging_p2p[i] / torque_mean

                        main_params[i].extend([cogging_p2p_percent, torque_mean, ripple_p2p, ripple_p2p_percent])

                        if len(amplitudes) < 20:
                            for _ in range(len(amplitudes), 20):
                                amplitudes.append("-")
                                phases.append("-")
                        elif len(amplitudes) > 20:
                            amplitudes = amplitudes[:20]  # max 20 harmonics
                            phases = phases[:20]

                        ripple_amplitudes[i].extend(amplitudes)
                        ripple_phases[i].extend(phases)

        return main_params, ripple_amplitudes, ripple_phases
