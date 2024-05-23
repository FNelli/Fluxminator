import os
from os import path
import itertools
import threading

import openpyxl

from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QFileDialog
from PyQt5.QtGui import QPixmap

from fluxminator.Model import Range, Parameter
from fluxminator.View import ParameterSetCreator
from fluxminator.Runner import Runner
from fluxminator.Constants import *

import fluxminator.support_functions as sup_fun
import ui.CustomWidgets as custom_widgets


class Presenter(Runner):
    """ MVP - Presenter class """

    def __init__(self, model, view):

        self.model = model      # MVC / Model
        self.view = view        # MVC / View

        self.flux_application_path = None

        self._interactor()
        self._interactor_parameter_set_creator()

        self._reset()

    def change_flux_directory(self):

        direction = str(QFileDialog.getExistingDirectory(self.view, "Select Flux directory!"))
        if direction == "":
            return

        if not os.path.exists(direction + FLUX_EXE_PATH):
            sup_fun.popup_message(self.view, "Invalid direction!")
            return

        self.flux_application_path = direction

        try:
            file = open(FLUX_PATH_INFO_FILE, 'w')
            file.write(direction)
            file.close()
        except IOError:
            sup_fun.popup_message(self.view, "Something went wrong; can't write the path information in the "
                                             "flux_directory.txt file")

    def _interactor(self):
        """ Bindings for functions to refresh the model data & the view based on user input in the main window GUI """

        """ Radio buttons for session selection """
        self.view.newSessionButton.setChecked(True)
        self.view.newSessionButton.clicked.connect(lambda: self._session_selector(new=True, old=False, summary=False))
        self.view.oldSessionButton.clicked.connect(lambda: self._session_selector(new=False, old=True, summary=False))
        self.view.summarySessionButton.clicked.connect(lambda: self._session_selector(new=False, old=False, summary=True))
        
        """ .FLU model selector """
        self.view.fluxModelButton.clicked.connect(self._flux_model_selector)
        
        """ Slot & Pole numbers """
        self.view.slotNumSpinBox.valueChanged.connect(lambda: self._set_motor_slot_number(reset=False))
        self.view.poleNumSpinBox.valueChanged.connect(lambda: self._set_motor_pole_number(reset=False))

        """ Display a pop-up window for creating the parameter set file """
        self.view.motorParamExcelButton.clicked.connect(self._show_parameter_set_creator)

        """ Scenario parameters """
        self.view.coggingScenarioCheckBox.setChecked(True)
        self.view.coggingScenarioCheckBox.stateChanged.connect(lambda: self._set_cogging_scenario(reset=False))
        
        self.view.rippleScenarioCheckBox.setChecked(True)
        self.view.rippleScenarioCheckBox.stateChanged.connect(lambda: self._set_ripple_scenario(reset=False))
        
        self.view.loadDepCheckBox.stateChanged.connect(lambda: self._set_load_dependency(reset=False))
        self.view.gammaDepCheckBox.stateChanged.connect(lambda: self._set_gamma_dependency(reset=False))
        
        self.view.gammaNameLineEdit.textChanged.connect(lambda: self._set_gamma_name(reset=False))
        self.view.currentNameLineEdit.textChanged.connect(lambda: self._set_current_name(reset=False))
        self.view.currentSourceNameLineEdit.textChanged.connect(lambda: self._set_current_source_name(reset=False))
        
        self.view.coggingMin.valueChanged.connect(lambda: self._set_cogging_position_range(sender="min"))
        self.view.coggingMax.valueChanged.connect(lambda: self._set_cogging_position_range(sender="max"))
        self.view.coggingStep.valueChanged.connect(lambda: self._set_cogging_position_range(sender="step"))

        self.view.rippleMin.valueChanged.connect(lambda: self._set_ripple_position_range(sender="min"))
        self.view.rippleMax.valueChanged.connect(lambda: self._set_ripple_position_range(sender="max"))
        self.view.rippleStep.valueChanged.connect(lambda: self._set_ripple_position_range(sender="step"))

        self.view.currentMin.valueChanged.connect(lambda: self._set_current_value_range(sender="min"))
        self.view.currentMax.valueChanged.connect(lambda: self._set_current_value_range(sender="max"))
        self.view.currentStep.valueChanged.connect(lambda: self._set_current_value_range(sender="step"))

        self.view.gammaMin.valueChanged.connect(lambda: self._set_gamma_value_range(sender="min"))
        self.view.gammaMax.valueChanged.connect(lambda: self._set_gamma_value_range(sender="max"))
        self.view.gammaStep.valueChanged.connect(lambda: self._set_gamma_value_range(sender="step"))

        """ Runner """
        self.view.runButton.clicked.connect(self._start_or_stop)
        self.view.clearButton.clicked.connect(self._reset)

        """ Extra settings """
        self.view.deleteCheckBox.setChecked(True)
        self.view.deleteCheckBox.stateChanged.connect(self._set_delete_solutions)
        
    def _interactor_parameter_set_creator(self):
        """ Bindings for functions to refresh the model & view based on user input in the parameter set creator GUI """

        # Add new parameter to the parameter set:
        self.view.parameterSetCreatorWidget.newP_addButton.clicked.connect(self._refresh_parameter_set)

        # Edit parameter values from the list widget:
        self.view.parameterSetCreatorWidget.pSet_listWidget.itemDoubleClicked.connect(self._edit_parameter)

        # Add custom parameter combinations to the parameter set:
        self.view.parameterSetCreatorWidget.customRow_addButton.clicked.connect(self._show_special_set_creator)

        # Set estimated time of all simulations:
        self.view.parameterSetCreatorWidget.create_runtimeSpinbox.valueChanged.connect(self._set_estimated_time)

        # Set final parameter set & create excel file:
        self.view.parameterSetCreatorWidget.create_excelCreatorButton.clicked.connect(self._create_parameter_table)

        # Clear:
        self.view.parameterSetCreatorWidget.newP_clearButton.clicked.connect(self._clear_parameter_editor_widget)

        # Enable the main GUI (called after closing the parameter set creator widget):
        self.view.parameterSetCreatorWidget.signal.connect(self._enable_main_gui)

    def _interactor_special_parameter_set_creator(self):
        """ Bindings for functions to refresh the model data & view based on user input in the special parameter
        combination creator widget of the GUI """

        # Add new combination to the parameter set:
        self.view.specialParameterSetCreatorWidget.addButton.clicked.connect(self._add_special_parameter_set)

        # Enable the parameter set creator GUI (called after closing the special parameter set creator widget):
        self.view.specialParameterSetCreatorWidget.signal.connect(self._enable_parameter_set_creator_gui)

    def _session_selector(self, new, old, summary):
        """ Session selector, reset the model data & and the view """
        
        self.model.sessionNew = new
        self.model.sessionOld = old
        self.model.sessionSummary = summary
        
        self._reset()

    def _flux_model_selector(self):
        """ Select a .FLU model for the simulation and update both the model data and the view """

        direction = str(QFileDialog.getExistingDirectory(self.view, "Select Flux model!"))

        if direction == "":
            return

        if direction[-4:] != '.FLU':
            sup_fun.popup_message(self.view, "Flux model path is invalid!")
            return
        
        path_data = direction.rsplit('/', 1)

        # To continue the simulations, the Project_Info and Sets files are necessary:
        if self.model.sessionOld or self.model.sessionSummary:
            if not os.path.exists(path_data[0] + "/Results/Project_Info.xlsx"):
                sup_fun.popup_message(self.view, "Results/Project_Info.xlsx is missing!")
                return
            if not os.path.exists(path_data[0] + "/Results/Sets.xlsx"):
                sup_fun.popup_message(self.view, "Results/Sets.xlsx is missing!")
                return
                
        """ We reach this point: valid .FLU model was choosen """

        # Model changes:
        self.model.fluxModel.set_modelPath(path_data[0])
        self.model.fluxModel.set_modelName(path_data[1])
        
        self.model.set_simulationTimes([])              # old time data would poop on the new time data
        self.model.set_progressState(0)
        
        # GUI changes:
        self.view.fluxModelLabel.setText(direction)

        self.view.progressBar.setValue(0)
        self.view.progressLabel.setText("Calculation: 0/0, estimated finish time: -")

        if path.exists(direction + '/preview1.png'):
            self.view.pixLabel.setPixmap(QPixmap(direction + '/preview1.png').scaled(300, 300, Qt.KeepAspectRatio))
        else:
            self.view.pixLabel.setPixmap(self.view.pixMap)

        if self.model.is_sessionNew():
            self.view.motorParamExcelButton.setDisabled(False)
            self.view.motorParamExcelButton.setToolTip("")
        else:
            self._open_existing_project()               # gather data to update the model and view
    
    def _open_existing_project(self):
        """ In the event of continuing a simulation already in progress or creating only a summary file,
        data from both the project info file and the parameter set file are necessary to update the model,
        the view, and commence the execution.
        The creation of the channel data template is also conducted here, without the specific parameter
        combination ID and values, which are always filled out before the upcoming simulation starts """

        try:
            model_path = self.model.fluxModel.get_modelPath()
            model_name = self.model.fluxModel.get_modelName()

            self.model.fluxModel.motor.set_parameterSetExcelPath(model_path + "/Results/Sets.xlsx")

            wb = openpyxl.load_workbook(model_path + "/Results/Project_Info.xlsx")
            ws = wb.worksheets[0]

            # MOTOR:
            slot_number = ws.cell(row=2, column=2).value
            pole_number = ws.cell(row=3, column=2).value

            self._set_motor_slot_number(value=slot_number)
            self._set_motor_pole_number(value=pole_number)

            # SCENARIO:
            scenario_id = ws.cell(row=4, column=2).value

            delete_results = ws.cell(row=15, column=2).value
            self.view.deleteCheckBox.setChecked(delete_results)

            # COGGING:
            if scenario_id == "2":
                cogging_scenario = False
                c_min, c_max, c_step = 0, 0, 0
            else:
                cogging_scenario = True
                c_values = ws.cell(row=5, column=2).value.split(" ")
                c_min, c_max, c_step = float(c_values[0]), float(c_values[1]), float(c_values[2])

            self.view.coggingScenarioCheckBox.setChecked(cogging_scenario)
            self.view.coggingMin.setValue(c_min)
            self.view.coggingMax.setValue(c_max)
            self.view.coggingStep.setValue(c_step)

            # RIPPLE:
            if scenario_id == "1":
                ripple_scenario = False
                r_min, r_max, r_step = 0, 0, 0
                cu_min, cu_max, cu_step = 0, 0, 0
                g_min, g_max, g_step = 0, 0, 0
            else:
                ripple_scenario = True
                r_values = ws.cell(row=6, column=2).value.split(" ")
                cu_values = ws.cell(row=8, column=2).value.split(" ")
                g_values = ws.cell(row=9, column=2).value.split(" ")

                r_min, r_max, r_step = float(r_values[0]), float(r_values[1]), float(r_values[2])
                cu_min, cu_max, cu_step = float(cu_values[0]), float(cu_values[1]), float(cu_values[2])
                g_min, g_max, g_step = float(g_values[0]), float(g_values[1]), float(g_values[2])

            self.view.rippleScenarioCheckBox.setChecked(ripple_scenario)
            self.view.rippleMin.setValue(r_min)
            self.view.rippleMax.setValue(r_max)
            self.view.rippleStep.setValue(r_step)

            self.view.currentMin.setValue(cu_min)
            self.view.currentMax.setValue(cu_max)
            self.view.currentStep.setValue(cu_step)

            self.view.gammaMin.setValue(g_min)
            self.view.gammaMax.setValue(g_max)
            self.view.gammaStep.setValue(g_step)

            # GAMMA & CURRENT:
            ripple_type = ws.cell(row=7, column=2).value
            gamma_dependency = ws.cell(row=10, column=2).value
            current_dependency = ws.cell(row=11, column=2).value

            gamma_name = ws.cell(row=12, column=2).value
            current_name = ws.cell(row=13, column=2).value
            current_source_name = ws.cell(row=14, column=2).value

            self.view.loadDepCheckBox.setChecked(gamma_dependency)
            self.view.gammaDepCheckBox.setChecked(current_dependency)

            self.view.gammaNameLineEdit.setText(gamma_name)
            self.view.currentNameLineEdit.setText(current_name)
            self.view.currentSourceNameLineEdit.setText(current_source_name)

            wb.close()

            self._disable()

            # CHANNEL DATA:
            # placeholders: {id} -> No-000X; {params} -> values for current No-000X

            data_model_path = model_path.replace('/', '\\\\') + "\n"
            data_model_name = model_name[:-4] + "\n"
            data_simulation_id = "{id}\n"               # exact value later for every simulation
            data_scenario = scenario_id + "\n"
            data_ripple_type = ripple_type + "\n"
            data_cogging = str(c_min) + " " + str(c_max) + " " + str(c_step) + "\n"
            data_ripple = str(r_min) + " " + str(r_max) + " " + str(r_step) + "\n"
            data_current = current_name + " " + str(cu_min) + " " + str(cu_max) + " " + str(cu_step) + "\n"
            data_gamma = gamma_name + " " + str(g_min) + " " + str(g_max) + " " + str(g_step) + "\n"
            data_parameters = "{params}\n"              # exact values later for every simulation
            data_delete_results = str(delete_results) + "\n"
            data_current_source = current_source_name + "\n"

            # Price calculation dependent data - TODO FUTURE DEVELOPMENT
            data_for_price_calculation = "Tooth\nRotor\nMagnet\nNone\n"

            channel_data_string = data_model_path + data_model_name + data_simulation_id \
                                  + data_scenario + data_ripple_type + data_cogging + data_ripple \
                                  + data_current + data_gamma + data_parameters + data_delete_results \
                                  + data_current_source + data_for_price_calculation

            self.model.set_channelData(channel_data_string)

        except IOError:
            sup_fun.popup_message(self.view, "Problem with Results/Project_Info.xlsx file!")
            self._reset()

    # FUNCTIONS TO SYNCHRONIZE MODEL & VIEW:

    def _set_cogging_scenario(self, reset=False):
        """ Set whether the cogging scenario is chosen or not """

        if reset:
            self.model.set_scenarioCogging(True)
            self.view.coggingScenarioCheckBox.setDisabled(False)
            self.view.coggingScenarioCheckBox.setChecked(True)
            disable = False
        else:
            is_cogging_scenario = self.view.coggingScenarioCheckBox.isChecked()
            self.model.set_scenarioCogging(is_cogging_scenario)
            disable = not is_cogging_scenario

        # If the cogging scenario is chosen, the cogging-related data should be editable on the GUI
        self.view.coggingMin.setDisabled(disable)
        self.view.coggingMax.setDisabled(disable)
        self.view.coggingStep.setDisabled(disable)    
        
    def _set_ripple_scenario(self, reset=False):
        """ Set whether the ripple scenario is chosen or not """

        if reset:
            self.model.set_scenarioRipple(True)
            self.view.rippleScenarioCheckBox.setDisabled(False)
            self.view.rippleScenarioCheckBox.setChecked(True)
            disable = False
        else:
            is_ripple_scenario = self.view.rippleScenarioCheckBox.isChecked()
            self.model.set_scenarioRipple(is_ripple_scenario)
            disable = not is_ripple_scenario

        # If the ripple scenario is chosen, the ripple-related data should be editable on the GUI
        self.view.loadDepCheckBox.setDisabled(disable)
        self.view.gammaDepCheckBox.setDisabled(disable)
        
        for i, j in itertools.product(range(1, 4), range(1, 4)):        # ripple, gamma, load spin boxes
            self.view.scenarioParameters[i][j].setDisabled(disable)
        
    def _set_load_dependency(self, reset=False):
        """ Set whether load dependency for the ripple scenario is chosen or not """
        if reset:
            self.model.set_dependencyLoad(False)
            self.view.loadDepCheckBox.setChecked(False)
        else:
            self.model.set_dependencyLoad(self.view.loadDepCheckBox.isChecked())
        
    def _set_gamma_dependency(self, reset=False):
        """ Set whether gamma dependency for the ripple scenario is chosen or not"""
        if reset:
            self.model.set_dependencyGamma(False)
            self.view.gammaDepCheckBox.setChecked(False)
        else:
            self.model.set_dependencyGamma(self.view.gammaDepCheckBox.isChecked())
    
    def _set_gamma_name(self, reset=False):
        """ Set the name of the gamma parameter for the Flux application """
        if reset:
            self.model.fluxModel.set_nameGamma("GAMMA")
            self.view.gammaNameLineEdit.setDisabled(False)
            self.view.gammaNameLineEdit.setText("GAMMA")
        else:
            self.model.fluxModel.set_nameGamma(self.view.gammaNameLineEdit.text())
            
    def _set_current_name(self, reset=False):
        """ Set the name of the current parameter for the Flux application """
        if reset:
            self.model.fluxModel.set_nameCurrent("IMAX")
            self.view.currentNameLineEdit.setDisabled(False)
            self.view.currentNameLineEdit.setText("IMAX")
        else:
            self.model.fluxModel.set_nameCurrent(self.view.currentNameLineEdit.getText())
            
    def _set_current_source_name(self, reset=False):
        """ Set the name of the current source parameter for the Flux application """
        if reset:
            self.model.fluxModel.set_nameCurrentSource("IPH_U")
            self.view.currentSourceNameLineEdit.setDisabled(False)
            self.view.currentSourceNameLineEdit.setText("IPH_U")
        else:
            self.model.fluxModel.set_nameCurrentSource(self.view.currentSourceNameLineEdit.getText())
            
    def _set_cogging_position_range(self, sender=None, reset=False):
        """ Set the rotor position for the cogging calculation """
        if reset:
            self.model.fluxModel.positionCogging.set_min(0)
            self.model.fluxModel.positionCogging.set_max(0)
            self.model.fluxModel.positionCogging.set_step(0)

            self.view.coggingMin.setValue(0)
            self.view.coggingMax.setValue(0)
            self.view.coggingStep.setValue(0)

        elif sender == "min":
            new_value = self.view.coggingMin.value()
            self.model.fluxModel.positionCogging.set_min(new_value)
        elif sender == "max":
            new_value = self.view.coggingMax.value()
            self.model.fluxModel.positionCogging.set_max(new_value)
        else:   # sender == "step"
            new_value = self.view.coggingStep.value()
            self.model.fluxModel.positionCogging.set_step(new_value)
        
    def _set_ripple_position_range(self, sender=None, reset=False):
        """ Set the rotor position for the ripple calculation """
        if reset:
            self.model.fluxModel.positionRipple.set_min(0)
            self.model.fluxModel.positionRipple.set_max(0)
            self.model.fluxModel.positionRipple.set_step(0)

            self.view.rippleMin.setValue(0)
            self.view.rippleMax.setValue(0)
            self.view.rippleStep.setValue(0)

        elif sender == "min":
            new_value = self.view.rippleMin.value()
            self.model.fluxModel.positionRipple.set_min(new_value)
        elif sender == "max":
            new_value = self.view.rippleMax.value()
            self.model.fluxModel.positionRipple.set_max(new_value)
        else:  # sender == "step"
            new_value = self.view.rippleStep.value()
            self.model.fluxModel.positionRipple.set_step(new_value)
        
    def _set_current_value_range(self, sender=None, reset=False):
        """ Set the current range for the simulation [A] """
        if reset:
            self.model.fluxModel.rangeCurrent.set_min(0)
            self.model.fluxModel.rangeCurrent.set_max(0)
            self.model.fluxModel.rangeCurrent.set_step(0)

            self.view.currentMin.setValue(0)
            self.view.currentMax.setValue(0)
            self.view.currentStep.setValue(0)

        elif sender == "min":
            new_value = self.view.currentMin.value()
            self.model.fluxModel.rangeCurrent.set_min(new_value)
        elif sender == "max":
            new_value = self.view.currentMax.value()
            self.model.fluxModel.rangeCurrent.set_max(new_value)
        else:  # sender == "step"
            new_value = self.view.currentStep.value()
            self.model.fluxModel.rangeCurrent.set_step(new_value)
        
    def _set_gamma_value_range(self, sender=None, reset=False):
        """ Set the gamma range for the simulation [eDeg] """
        if reset:
            self.model.fluxModel.rangeGamma.set_min(0)
            self.model.fluxModel.rangeGamma.set_max(0)
            self.model.fluxModel.rangeGamma.set_step(0)

            self.view.gammaMin.setValue(0)
            self.view.gammaMax.setValue(0)
            self.view.gammaStep.setValue(0)

        elif sender == "min":
            new_value = self.view.gammaMin.value()
            self.model.fluxModel.rangeGamma.set_min(new_value)
        elif sender == "max":
            new_value = self.view.gammaMax.value()
            self.model.fluxModel.rangeGamma.set_max(new_value)
        else:  # sender == "step"
            new_value = self.view.gammaStep.value()
            self.model.fluxModel.rangeGamma.set_step(new_value)
    
    def _set_motor_slot_number(self, reset=False, value=None):
        """ Set the slot number of the motor """
        if reset:
            self.model.fluxModel.motor.set_numberSlot(0)
            self.view.slotNumSpinBox.setDisabled(False)
            self.view.slotNumSpinBox.setValue(0)
        elif value is not None:
            self.model.fluxModel.motor.set_numberSlot(value)
            self.view.slotNumSpinBox.setValue(value)
        else:
            self.model.fluxModel.motor.set_numberSlot(self.view.slotNumSpinBox.value())
            
    def _set_motor_pole_number(self, reset=False, value=None):
        """ Set the pole number of the motor """
        if reset:
            self.model.fluxModel.motor.set_numberPole(0)
            self.view.poleNumSpinBox.setDisabled(False)
            self.view.poleNumSpinBox.setValue(0)
        elif value is not None:
            self.model.fluxModel.motor.set_numberPole(value)
            self.view.poleNumSpinBox.setValue(value)
        else:
            self.model.fluxModel.motor.set_numberPole(self.view.poleNumSpinBox.value())
            
    def _set_delete_solutions(self, reset=False):
        """ Set whether raw results from the simulations should be deleted or not """
        if reset:
            self.model.set_deleteScenarioSolutions(True)
            self.view.deleteCheckBox.setDisabled(False)
            self.view.deleteCheckBox.setChecked(True)
        else:
            self.model.set_deleteScenarioSolutions(self.view.deleteCheckBox.isChecked())

    # FUNCTIONS FOR THE PARAMETER SETS FILE CREATOR:

    def _show_parameter_set_creator(self):
        """ Show the parameter set creator widget for generating the Sets.xlsx file for the simulations """
        self.view.parameterSetCreatorWidget.show()

        # Temporarily disable the main GUI to avoid inconsistency
        self.view.setEnabled(False)

    def _refresh_parameter_set(self):
        """ Refresh parameter set: Add new parameter / Edit existing parameter """

        parameter_type = self.view.parameterSetCreatorWidget.newP_typeCombo.currentText()
        parameter_name = self.view.parameterSetCreatorWidget.newP_nameLineEdit.text()
        parameter_desc = self.view.parameterSetCreatorWidget.newP_descLineEdit.text()

        if parameter_name == "":
            sup_fun.popup_message(self.view.parameterSetCreatorWidget, "Name is missing!")
            return

        parameter_min = self.view.parameterSetCreatorWidget.newP_minSpinbox.value()
        parameter_max = self.view.parameterSetCreatorWidget.newP_maxSpinbox.value()
        parameter_step = self.view.parameterSetCreatorWidget.newP_stepSpinbox.value()

        parameter_range = Range(min_=parameter_min, max_=parameter_max, step=parameter_step)

        # Check whether the parameter value range is valid or not:
        if not parameter_range.is_valid_range():
            sup_fun.popup_message(self.view.parameterSetCreatorWidget, "Invalid range!")
            return

        # Check whether the parameter name is unique or not:
        if self.view.parameterSetCreatorWidget.newP_addButton.text() == "Add":
            if self.model.fluxModel.motor.is_existing_parameter(parameter_name):
                sup_fun.popup_message(self.view.parameterSetCreatorWidget, "The parameter already exist.")
                return

        parameter = Parameter(type_=parameter_type, name=parameter_name, desc=parameter_desc, range_=parameter_range)

        # Add new parameter set to the parameter set list:
        if self.view.parameterSetCreatorWidget.newP_addButton.text() == "Add":

            # Model update (update the special combination list too if it exists):
            self.model.fluxModel.motor.add_parameter(parameter)
            self.model.fluxModel.motor.update_special_parameter_set(action="Add", new_name=parameter_name,
                                                                    default_value=parameter_min)
            # View update:
            self.view.parameterSetCreatorWidget.pSet_listWidget.addItem(parameter_name)

            # Update the row number (number of parameter set combinations) displayed on the GUI:
            self._update_row_number()

        # Edit a parameter set from the parameter set list:
        else:   # "Overwrite"
            list_widget_parameter_item = self.view.parameterSetCreatorWidget.pSet_listWidget.currentItem()

            orig_name = list_widget_parameter_item.text()

            # Model update (update the special combination list too if it exists):
            self.model.fluxModel.motor.update_parameter(orig_name=orig_name, new_parameter=parameter)
            self.model.fluxModel.motor.update_special_parameter_set(action="Edit", orig_name=orig_name,
                                                                    new_name=parameter_name, default_value=parameter_min)
            # View update:
            list_widget_parameter_item.setText(parameter_name)

            # Update the row number (number of parameter set combinations) displayed on the GUI:
            self._update_row_number()

        # Clear the GUI:
        self._clear_parameter_editor_widget()

    def _edit_parameter(self):
        """ Edit parameter after double click on a listWidget item """

        item = self.view.parameterSetCreatorWidget.pSet_listWidget.currentItem()  # the item that was double-clicked
        item_name = item.text()

        orig_parameter = self.model.fluxModel.motor.get_parameter_by_name(parameter_name=item_name)

        # Pop up a message box for choosing an action (edit/delete):
        result = self.view.parameterSetCreatorWidget.parameter_editor_message_box()

        # EDIT - load data to the parameter set editor part of the GUI:
        if result == 0:
            self.view.parameterSetCreatorWidget.newP_typeCombo.setCurrentText(orig_parameter.get_type())
            self.view.parameterSetCreatorWidget.newP_nameLineEdit.setText(orig_parameter.get_name())
            self.view.parameterSetCreatorWidget.newP_descLineEdit.setText(orig_parameter.get_desc())

            self.view.parameterSetCreatorWidget.newP_minSpinbox.setValue(orig_parameter.range.get_min())
            self.view.parameterSetCreatorWidget.newP_maxSpinbox.setValue(orig_parameter.range.get_max())
            self.view.parameterSetCreatorWidget.newP_stepSpinbox.setValue(orig_parameter.range.get_step())

            self.view.parameterSetCreatorWidget.newP_title.setText("Edit parameter")
            self.view.parameterSetCreatorWidget.newP_addButton.setText("Overwrite")
            self.view.parameterSetCreatorWidget.newP_clearButton.setText("Cancel")

        # DELETE:
        elif result == 1:
            # Model update (update the special combination list too if it exists):
            self.model.fluxModel.motor.delete_parameter(orig_parameter)
            self.model.fluxModel.motor.update_special_parameter_set(action="Delete", orig_name=item_name)

            # View update:
            self.view.parameterSetCreatorWidget.pSet_listWidget.takeItem(
                self.view.parameterSetCreatorWidget.pSet_listWidget.row(item))

            # Update the row number (number of parameter set combinations) displayed on the GUI:
            self._update_row_number()

    def _update_row_number(self):
        """ Update the row number (number of parameter set combinations) displayed on the GUI """
        parameters = self.model.fluxModel.motor.get_parameters()

        row_number = 0
        if len(parameters) > 0:
            row_number = 1
            for p in parameters:
                row_number *= p.range.get_numberOfSteps()
            row_number += self.model.fluxModel.motor.get_specialRowNumber()

        # View update:
        self.view.parameterSetCreatorWidget.create_rowNumberLabel.setText("Number of rows: %i" % row_number)
        self._set_estimated_time(row_number)

    def _set_estimated_time(self, row_number):
        """ Display the estimated time of all the simulations based on the number of the parameter combinations
        and the estimated time for one simulation """

        # Estimated time of one simulation:
        simulation_time = self.view.parameterSetCreatorWidget.create_runtimeSpinbox.value()

        if row_number == 0 or simulation_time == 0:
            self.view.parameterSetCreatorWidget.create_sumRuntimeLabel.setText("Estimated runtime: -")
        else:
            time_in_mins = row_number * simulation_time             # whole simulation time in minutes
            mins = time_in_mins % 60                                # remaining minutes above hours
            remaining_time_in_hours = (time_in_mins - mins) / 60    # all remaining time in hours
            hours = remaining_time_in_hours % 24                    # remaining hours above days
            days = (remaining_time_in_hours - hours) / 24           # days

            min_str = "min"
            hour_str = "hour"
            day_str = "day"

            if mins > 1:
                min_str += "s"
            if hours > 1:
                hour_str += "s"
            if days > 1:
                day_str += "s"

            time_text = "Estimated time: {0} {3}, {1} {4}, {2} {5}.".format(int(days), int(hours), int(mins),
                                                                            day_str, hour_str, min_str)
            self.view.parameterSetCreatorWidget.create_sumRuntimeLabel.setText(time_text)

    def _create_parameter_table(self):
        """ Create parameter set file via the functions of the Motor object """

        if len(self.model.fluxModel.motor.get_parameters()) == 0:
            sup_fun.popup_message(self.view.parameterSetCreatorWidget, "No parameters yet.")
            return

        # Parameter table (lists in list) with all parameter combinations, including special sets:
        parameter_table = self.model.fluxModel.motor.create_final_parameter_set()

        try:
            # Create the Sets.xlsx
            self.model.fluxModel.motor.create_parameter_table_excel(parameter_table, self.model.fluxModel.get_modelPath())
            self.view.parameterSetCreatorWidget.create_excelCreatorButton.setText("Overwrite")
        except IOError:
            sup_fun.popup_message(self.view.parameterSetCreatorWidget, "Something went wrong during saving "
                                                                       "the parameter sets file. Close the "
                                                                       "'Sets.xlsx' file if its open!")

    def _clear_parameter_editor_widget(self):
        """ Clear the parameter set editor part of the GUI """

        self.view.parameterSetCreatorWidget.newP_typeCombo.setCurrentText("GP")
        self.view.parameterSetCreatorWidget.newP_nameLineEdit.setText("")
        self.view.parameterSetCreatorWidget.newP_descLineEdit.setText("")

        self.view.parameterSetCreatorWidget.newP_minSpinbox.setValue(0.00)
        self.view.parameterSetCreatorWidget.newP_maxSpinbox.setValue(0.00)
        self.view.parameterSetCreatorWidget.newP_stepSpinbox.setValue(0.00)

        self.view.parameterSetCreatorWidget.newP_title.setText("Add new parameter")
        self.view.parameterSetCreatorWidget.newP_addButton.setText("Add")
        self.view.parameterSetCreatorWidget.newP_clearButton.setText("Clear")

    def _enable_main_gui(self):
        """ Enable the main GUI (called after closing the parameter set creator widget) """
        self.view.setEnabled(True)

    # FUNCTIONS FOR THE SPECIAL PARAMETER SET CREATOR:

    def _show_special_set_creator(self):
        """ After pressing the 'New' button on the 'Add special parameter combinations' section of the parameter set
        creator GUI, a widget will pop up for adding a new combination """

        # Always create a new widget with the existing parameter set list
        motor_parameters = self.model.fluxModel.motor.get_parameters()
        self.view.specialParameterSetCreatorWidget = custom_widgets.SpecialParameterSetCreator(motor_parameters)

        # Bindings between the View & the Model
        self._interactor_special_parameter_set_creator()

        self.view.specialParameterSetCreatorWidget.show()

        # Temporarily disable the parameter set creator widget to avoid inconsistency
        self.view.parameterSetCreatorWidget.setEnabled(False)

    def _add_special_parameter_set(self):
        """ Add a special parameter combination to the parameter table """

        # If there are no parameters yet, special combinations cannot be added:
        if len(self.model.fluxModel.motor.get_parameters()) == 0:
            return

        # Retrieve the parameter names and their corresponding values from the view:
        parameter_labels = self.view.specialParameterSetCreatorWidget.get_parameter_labels()
        value_spinboxes = self.view.specialParameterSetCreatorWidget.get_parameter_value_containers()

        names = []
        values = []

        for i in range(len(parameter_labels)):
            names.append(parameter_labels[i].text())
            values.append(value_spinboxes[i].value())

        # Model update:
        self.model.fluxModel.motor.add_special_parameter_combination(parameter_names=names, parameter_values=values)

        # View update:
        self.view.specialParameterSetCreatorWidget.reset_background()

        # Update the row number (number of parameter set combinations) displayed on the GUI:
        self._update_row_number()

    def _enable_parameter_set_creator_gui(self):
        """ Enable the parameter set creator GUI (called after closing the special combination creator widget) """
        self.view.parameterSetCreatorWidget.setEnabled(True)

    # FUNCTIONS FOR RUNNING:

    def _start_or_stop(self):
        """ This function is called whenever the user presses the RUN/STOP button on the main GUI """

        # RUN:
        if not self.model.get_executionInProgress():

            if not self.view.runButton.text() == "Continue":
                # Check the validity of the settings:
                if (not self._check_flux_application()
                        or not self._check_motor_parameters()
                        or not self._check_scenario_parameters()):
                    return

                # Project Info file & Channel data:
                if self.model.is_sessionNew():
                    try:
                        self.model.create_info_files()
                    except IOError:
                        sup_fun.popup_message(self.view, "There is a problem with the Project_Info file.")
                        return

            # Get gamma values for 2Stack & 3Stack:
            try:
                self.model.fluxModel.read_skewing_file()
            except IOError:
                sup_fun.popup_message(self.view, "There is a problem with the Skewing file.")

            self.model.set_executionInProgress(True)

            if not self.model.is_sessionSummary():
                self.view.runButton.setText("Stop")
                self.view.runButton.setToolTip("Stop at the end of the current simulation.")

            # Disable most of the widgets on the main GUI during execution to avoid inconsistency:
            self._disable()
            self.view.clearButton.setDisabled(True)

            # STARTING A NEW THREAD FOR FLUX:
            flux_thread = threading.Thread(target=self._execution(), args=())
            flux_thread.start()

        # STOP:
        else:
            # Message box occurs when the stop button is pushed:
            result = self.view.stop_simulation_message_box()

            if result == 0:
                self.model.set_executionInProgress(False)
                self.view.runButton.setText("Wait")
                self.view.runButton.setDisabled(True)
                self.view.runButton.setToolTip("Waiting for the current calculation to finish.")

    def _check_flux_application(self):
        """ Check if the application path for the Flux software is valid in the 'flux_directory.txt' file """

        try:
            flux_application_path_file = open(FLUX_PATH_INFO_FILE, 'r')
            flux_application_path = flux_application_path_file.readline()
            flux_application_path_file.close()

            if flux_application_path.strip() == "":
                sup_fun.popup_message(self.view, "Flux application path is missing!")
                return False

            if not os.path.exists(flux_application_path + FLUX_EXE_PATH):
                sup_fun.popup_message(self.view, "Flux application path is invalid!")
                return False

            self.flux_application_path = flux_application_path

        except IOError:
            sup_fun.popup_message(self.view, "Problem with flux_directory.txt!")
            return False

        return True

    def _check_motor_parameters(self):
        """ Check for a valid motor model, parameter set file, and slot & pole numbers before starting the execution """

        if self.model.fluxModel.get_modelPath() is None or self.model.fluxModel.get_modelName() is None:
            sup_fun.popup_message(self.view, "Flux model path is missing!")
            return False

        if self.model.fluxModel.motor.get_parameterSetExcelPath() == "":
            sup_fun.popup_message(self.view, "Results/Sets.xlsx file does not exist!")
            return False

        try:
            os.path.exists(self.model.fluxModel.motor.get_parameterSetExcelPath())
        except IOError:
            sup_fun.popup_message(self.view, "Results/Sets.xlsx file does not exist!")
            return False

        if self.model.fluxModel.motor.get_numberSlot() == 0:
            sup_fun.popup_message(self.view, "Slot number can't be zero!")
            return False

        if self.model.fluxModel.motor.get_numberPole() == 0:
            sup_fun.popup_message(self.view, "Pole number can't be zero!")
            return False

        return True

    def _check_scenario_parameters(self):
        """ Check if the scenario parameters are valid before execution """

        # Cogging & Ripple scenario
        if not self.model.is_scenarioCogging() and not self.model.is_scenarioRipple():
            sup_fun.popup_message(self.view, "You have to choose at least one scenario!")
            return False

        if self.model.is_scenarioCogging():
            if (not self.model.fluxModel.positionCogging.is_valid_range()
                    or self.model.fluxModel.positionCogging.get_step() == 0):
                sup_fun.popup_message(self.view, "Invalid rotor position values for cogging scenario!")
                return False

        if self.model.is_scenarioRipple():
            if (not self.model.fluxModel.positionRipple.is_valid_range()
                    or self.model.fluxModel.positionRipple.get_step() == 0):
                sup_fun.popup_message(self.view, "Invalid rotor position values for ripple scenario!")
                return False

            if not self.model.fluxModel.rangeGamma.is_valid_range():
                sup_fun.popup_message(self.view, "Invalid gamma values for ripple scenario!")
                return False

            if (not self.model.fluxModel.rangeCurrent.is_valid_range()
                    or self.model.fluxModel.rangeCurrent.get_max() == 0):
                sup_fun.popup_message(self.view, "Invalid current values for ripple scenario!")
                return False

            if self.model.is_dependencyGamma() and self.model.fluxModel.rangeGamma.get_step() == 0:
                sup_fun.popup_message(self.view, "Gamma dependency sheet has been requested, "
                                                 "but there is only one variation of gamma!")
                return False

            if self.model.is_dependencyLoad() and self.model.fluxModel.rangeCurrent.get_step() == 0:
                sup_fun.popup_message(self.view, "Load dependency sheet has been requested, "
                                                 "but there is only one variation of current!")
                return False

        # Gamma & Current names for Flux
        if self.model.fluxModel.get_nameGamma() == "":
            sup_fun.popup_message(self.view, "'Gamma name' is missing!")
            return False

        if self.model.fluxModel.get_nameCurrent() == "":
            sup_fun.popup_message(self.view, "'Current name' is missing!")
            return False

        if self.model.fluxModel.get_nameCurrentSource() == "":
            sup_fun.popup_message(self.view, "'Current source name' is missing!")
            return False

        return True

    def _execution(self):
        """ Start the execution (simulations or summary file creation), and after that, perform cleanup based on
        the result of the execution """

        execution_result = self._run(self)
        print("EXECUTION RESULT:", execution_result)
        print("NEXT SIMULATION ID:", self.model.get_simulationID())

        # After the execution is complete:
        self.model.set_executionInProgress(False)
        self.view.runButton.setToolTip("")
        self.view.runButton.setDisabled(False)
        self.view.clearButton.setDisabled(False)

        if execution_result == EXECUTION_DONE or execution_result == MISSING_RAW_FILES:
            self.view.runButton.setText("Run")
            self._disable(False)

            if execution_result == MISSING_RAW_FILES:
                sup_fun.popup_message(self.view, MISSING_RAW_FILES)
                Presenter._set_progress(self, reset=True)                       # defined by the Runner object
            else:
                sup_fun.popup_message(view=self.view, message=EXECUTION_DONE, title="Information")

            return

        self.view.runButton.setText("Continue")

        if execution_result == EXECUTION_STOP:
            return

        if execution_result == RESULT_FILE_IO_ERROR:
            self.model.set_summarySavingProblem(True)
            message = RESULT_FILE_IO_ERROR + "\nPlease close the result files and press the 'Continue' button!"
            sup_fun.popup_message(self.view, message)

        # RAW_COGGING_FILE_ERROR, RAW_RIPPLE_FILE_ERROR, EXECUTION_ERROR:
        else:
            sup_fun.popup_message(self.view, execution_result)

    # CLEAR:

    def _disable(self, disable=True):
        """ Disable most of the widgets on the main GUI (during execution to avoid inconsistency) """

        self.view.newSessionButton.setDisabled(disable)
        self.view.oldSessionButton.setDisabled(disable)
        self.view.summarySessionButton.setDisabled(disable)

        self.view.deleteCheckBox.setDisabled(disable)

        self.view.gammaNameLineEdit.setDisabled(disable)
        self.view.currentNameLineEdit.setDisabled(disable)
        self.view.currentSourceNameLineEdit.setDisabled(disable)

        self.view.fluxModelButton.setDisabled(disable)

        self.view.slotNumSpinBox.setDisabled(disable)
        self.view.poleNumSpinBox.setDisabled(disable)

        self.view.motorParamExcelButton.setDisabled(disable)
        self.view.motorParamExcelButton.setToolTip("")

        self.view.coggingScenarioCheckBox.setDisabled(disable)
        self.view.rippleScenarioCheckBox.setDisabled(disable)

        self.view.loadDepCheckBox.setDisabled(disable)
        self.view.gammaDepCheckBox.setDisabled(disable)

        self.view.coggingScenarioCheckBox.setDisabled(disable)

        for i, j in itertools.product(range(0, 4), range(1, 4)):  # cogging, ripple, gamma, load spinboxes
            self.view.scenarioParameters[i][j].setDisabled(disable)

        # self.view.clearButton.setDisabled(disable)

    def _reset(self):
        """ Initialize model data and the view """

        # Model reset:

        self.model.set_executionInProgress(False)
        self.model.set_simulationID(1)
        self.model.set_channelData("")
        self.model.set_summarySavingProblem(False)
        self.model.set_projectInfoFilePath("")

        self.model.fluxModel.set_modelPath(None)
        self.model.fluxModel.set_modelName(None)

        self.model.fluxModel.set_gammasFor2Stack(None)
        self.model.fluxModel.set_gammasFor3Stack(None)

        self.model.fluxModel.motor.clear_parameter_set()

        # Model & GUI reset:

        self._set_gamma_name(reset=True)
        self._set_current_name(reset=True)
        self._set_current_source_name(reset=True)

        self._set_cogging_scenario(reset=True)
        self._set_ripple_scenario(reset=True)
        self._set_load_dependency(reset=True)
        self._set_gamma_dependency(reset=True)

        self._set_cogging_position_range(reset=True)
        self._set_ripple_position_range(reset=True)
        self._set_current_value_range(reset=True)
        self._set_gamma_value_range(reset=True)

        self._set_motor_slot_number(reset=True)
        self._set_motor_pole_number(reset=True)

        self._set_delete_solutions(reset=True)

        Presenter._set_progress(self, reset=True)                       # defined by the Runner object

        # GUI reset:

        self.view.fluxModelLabel.setText("Missing!")
        self.view.pixLabel.setPixmap(self.view.pixMap)

        self.view.motorParamExcelButton.setDisabled(True)
        self.view.motorParamExcelButton.setToolTip("Select Flux model before creating the parameter set file!")

        self.view.runButton.setDisabled(False)
        self.view.runButton.setText("Run")
        self.view.clearButton.setDisabled(False)

        self._disable(disable=False)

        self.view.parameterSetCreatorWidget = ParameterSetCreator()
        self._interactor_parameter_set_creator()


